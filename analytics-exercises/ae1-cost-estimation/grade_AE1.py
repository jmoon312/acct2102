"""
Grade a student's AE1 workbook.

Usage:
  python3 grade_AE1.py path/to/student.xlsx
  python3 grade_AE1.py path/to/folder/   # grades every .xlsx in folder

Reads the student's "Answer Sheet" tab (column D, rows 5-19) and compares
to expected values computed from the canonical Data tab (so the script
works even if a student accidentally edited Data — we recompute from
the canonical seed).

Prints a per-question report and a summary line. Writes a CSV next to
each student file with the per-question scores.
"""
import sys, os, csv
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook

SEED = 1590
def canonical_data():
    rng = np.random.default_rng(SEED)
    n = 24
    batches = rng.integers(60, 195, n)
    abs_ = rng.integers(300, 700, n)
    units = batches * abs_
    mh = 4.0 * batches + 0.012 * units + rng.normal(0, 80, n)
    dlh = 0.06 * units + rng.normal(0, 200, n)
    rm = 12 * units + rng.normal(0, 800, n)
    il = 20000 + 200 * batches + rng.normal(0, 1500, n)
    s = 10
    batches[s] = 35; units[s] = 35*480
    mh[s] = 4.0*35 + 0.012*units[s] + rng.normal(0, 80)
    dlh[s] = 0.06*units[s] + rng.normal(0, 200)
    rm[s] = 12*units[s] + rng.normal(0, 800)
    il[s] = 20000 + 200*35 + 3000
    return {
        'batches': batches.astype(int),
        'units': units.astype(int),
        'mh': np.round(mh).astype(int),
        'dlh': np.round(dlh).astype(int),
        'rm': np.round(rm).astype(int),
        'il': np.round(il).astype(int),
    }

D = canonical_data()
b, y = D['batches'], D['il']

slope, intercept = np.polyfit(b, y, 1)
r2 = np.corrcoef(b, y)[0,1]**2
corr = np.corrcoef(b, y)[0,1]

hi, lo = int(np.argmax(b)), int(np.argmin(b))
hl_vc = (y[hi] - y[lo]) / (b[hi] - b[lo])
hl_fc = y[hi] - hl_vc * b[hi]

new_fc = intercept * 1.20
new_vc = slope * 0.90
qstar = (new_fc - intercept) / (slope - new_vc)
cost_at_qstar = intercept + slope * qstar

# (qid, expected_value, kind, tol, points, label)
EXPECTED = [
    ("1.1", "batches",                  "text", None,  10, "Best cost driver"),
    ("1.2", corr,                       "abs",  0.005,  5, "Correlation r"),
    ("2.1", slope,                      "abs",  2.00,   8, "VC (slope)"),
    ("2.2", intercept,                  "abs",  300,    8, "FC (intercept)"),
    ("2.3", r2,                         "abs",  0.005,  5, "R²"),
    ("2.4", hl_vc,                      "abs",  2.00,   5, "HL VC"),
    ("2.5", hl_fc,                      "abs",  300,    5, "HL FC"),
    ("3.1", intercept + slope*75,       "abs",  300,    5, "Predicted at 75"),
    ("3.2", intercept + slope*175,      "abs",  300,    5, "Predicted at 175"),
    ("3.3", intercept + slope*250,      "abs",  300,    5, "Predicted at 250"),
    ("4.1", new_fc,                     "abs",  300,    5, "New FC"),
    ("4.2", new_vc,                     "abs",  2.00,   5, "New VC"),
    ("4.3", qstar,                      "abs",  2.0,   14, "Indifference Q*"),
    ("4.4", cost_at_qstar,              "abs",  300,    8, "Cost at Q*"),
    ("4.5", "expanded",                 "text", None,   7, "Cheaper at 240"),
]

ANSWER_ROW_MAP = {q[0]: 5 + i for i, q in enumerate(EXPECTED)}

def grade_one(path):
    wb = load_workbook(path, data_only=True)
    if "Answer Sheet" not in wb.sheetnames:
        print(f"[{path}] No 'Answer Sheet' tab — skipping.")
        return None
    ans = wb["Answer Sheet"]
    rows = []
    total_pts = 0
    max_pts = 0
    for qid, exp, kind, tol, pts, label in EXPECTED:
        r = ANSWER_ROW_MAP[qid]
        student_val = ans.cell(row=r, column=4).value
        max_pts += pts
        if kind == "text":
            stu_norm = ("" if student_val is None else str(student_val)).strip().lower()
            exp_norm = exp.strip().lower()
            passed = stu_norm == exp_norm
        else:
            try:
                stu_f = float(student_val)
                passed = abs(stu_f - exp) <= tol
            except (TypeError, ValueError):
                passed = False
                stu_f = student_val
        if passed:
            total_pts += pts
        rows.append({
            "qid": qid, "label": label, "expected": exp,
            "student": student_val, "passed": passed, "points": pts if passed else 0, "max": pts,
        })
    return {"path": path, "rows": rows, "total": total_pts, "max": max_pts}


def print_report(result):
    if result is None: return
    print(f"\n=== {os.path.basename(result['path'])} ===")
    print(f"{'#':<5}{'Question':<24}{'Expected':>16}{'Student':>16}  Pass?  Pts")
    print("-" * 72)
    for r in result["rows"]:
        exp = r["expected"]
        stu = r["student"]
        exp_s = exp if isinstance(exp, str) else f"{exp:>16,.4f}".rstrip("0").rstrip(".")
        stu_s = stu if isinstance(stu, str) else (f"{stu:>16,.4f}".rstrip("0").rstrip(".") if stu is not None else "—")
        mark = "PASS" if r["passed"] else "FAIL"
        print(f"{r['qid']:<5}{r['label']:<24}{str(exp_s):>16}{str(stu_s):>16}  {mark:<6}{r['points']}/{r['max']}")
    print(f"\nTOTAL: {result['total']} / {result['max']}  ({100*result['total']/result['max']:.1f}%)")


def write_csv(result, out_path):
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["qid","label","expected","student","passed","points","max"])
        for r in result["rows"]:
            w.writerow([r["qid"], r["label"], r["expected"], r["student"],
                        r["passed"], r["points"], r["max"]])
        w.writerow([])
        w.writerow(["TOTAL","","","", "", result["total"], result["max"]])

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    target = Path(sys.argv[1])
    files = []
    if target.is_dir():
        files = sorted(target.glob("*.xlsx"))
    elif target.exists():
        files = [target]
    else:
        print(f"Path not found: {target}"); sys.exit(1)
    for f in files:
        result = grade_one(str(f))
        if result is None: continue
        print_report(result)
        write_csv(result, str(f.with_suffix(".grade.csv")))
        print(f"Wrote {f.with_suffix('.grade.csv').name}")
