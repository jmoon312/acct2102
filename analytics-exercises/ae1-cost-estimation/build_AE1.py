"""
Build AE1_Student.xlsx and AE1_Instructor.xlsx.
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# ---- Style palette ----
NAVY = "1A2744"
CREAM = "F4F0E8"
ORANGE = "D85A30"
INPUT_YELLOW = "FFF2CC"
GREEN_PASS = "C6EFCE"
RED_FAIL = "FFC7CE"
GRAY = "E7E6E6"
LIGHT_NAVY = "E8ECF2"

def F(size=11, bold=False, italic=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, italic=italic, color=color)

def fill(c): return PatternFill("solid", start_color=c, end_color=c)

thin = Side(border_style="thin", color="999999")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
al_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
al_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
al_r = Alignment(horizontal="right", vertical="center")
al_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ---- Data ----
SEED = 1590
def generate_data():
    rng = np.random.default_rng(SEED)
    n = 24
    batches = rng.integers(60, 195, n)
    abs_ = rng.integers(300, 700, n)
    units = batches * abs_
    mh = 4.0 * batches + 0.012 * units + rng.normal(0, 80, n)
    dlh = 0.06 * units + rng.normal(0, 200, n)
    rm = 12 * units + rng.normal(0, 800, n)
    il = 20000 + 200 * batches + rng.normal(0, 1500, n)
    s = 10
    batches[s] = 35; units[s] = 35*480
    mh[s] = 4.0*35 + 0.012*units[s] + rng.normal(0, 80)
    dlh[s] = 0.06*units[s] + rng.normal(0, 200)
    rm[s] = 12*units[s] + rng.normal(0, 800)
    il[s] = 20000 + 200*35 + 3000
    months = pd.date_range('2024-01-01', periods=n, freq='MS').strftime('%b-%Y').tolist()
    return pd.DataFrame({
        'Month': months,
        'Batches': batches.astype(int),
        'Units (cases)': units.astype(int),
        'Machine Hours': np.round(mh).astype(int),
        'Direct Labor Hours': np.round(dlh).astype(int),
        'Raw Material (lbs)': np.round(rm).astype(int),
        'Indirect Labor Cost': np.round(il).astype(int),
    })

DATA = generate_data()
N_ROWS = len(DATA)  # 24
DATA_FIRST_ROW = 4  # Excel row where data starts on Data tab (after header)
DATA_LAST_ROW = DATA_FIRST_ROW + N_ROWS - 1


def banner(ws, row, text, span=12, color=NAVY, white=True):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F(size=14, bold=True, color="FFFFFF" if white else NAVY)
    c.fill = fill(color)
    c.alignment = al_l
    ws.row_dimensions[row].height = 26

def section(ws, row, text, span=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F(bold=True, color=NAVY)
    c.fill = fill(CREAM)
    c.alignment = al_l
    ws.row_dimensions[row].height = 20

def write(ws, row, col, val, font=None, fillc=None, align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    if font is not None: c.font = font
    if fillc is not None: c.fill = fill(fillc)
    if align is not None: c.alignment = align
    if border is not None: c.border = border
    if fmt is not None: c.number_format = fmt
    return c


# ===========================================================================
# STUDENT WORKBOOK
# ===========================================================================
def build_student(path):
    wb = Workbook()
    # Default sheet → Instructions
    ws = wb.active
    ws.title = "Instructions"
    build_instructions(ws)
    build_data(wb.create_sheet("Data"))
    build_step1(wb.create_sheet("Step 1"))
    build_step2(wb.create_sheet("Step 2"))
    build_step3(wb.create_sheet("Step 3"))
    build_step4(wb.create_sheet("Step 4"))
    build_answer_sheet(wb.create_sheet("Answer Sheet"), is_instructor=False)
    wb.save(path)
    print(f"Saved {path}")


# ---- Instructions tab ----
def build_instructions(ws):
    ws.column_dimensions['A'].width = 2
    for col, w in zip("BCDEFGHI", [12]*8):
        ws.column_dimensions[col].width = w
    ws.column_dimensions['B'].width = 100  # main text column

    banner(ws, 1, "ACCT 2102 — Analytics Exercise 1: Cost Estimation", span=10)
    banner(ws, 2, "Pearson Specialty Foods, Inc.", span=10, color=ORANGE)

    blocks = [
        ("Setting",
         "You are a financial analyst at Pearson Specialty Foods, a contract food manufacturer that "
         "produces packaged snacks for grocery and CPG clients. Allergen controls require a full sanitation "
         "between every product run. The operations team has provided 24 months of historical data, and "
         "your manager wants you to build a model of monthly indirect labor cost so it can be predicted "
         "for next year's budget — and used to evaluate a possible plant expansion."),
        ("How to work this exercise",
         "Work through the four Step tabs in order, using the data on the Data tab. Each Step tab includes "
         "a short prompt, a scaffold for your work, and labeled cells (highlighted yellow) where your "
         "final answers should land. When you're finished, complete the Answer Sheet tab — each answer "
         "there should be a formula linking back to a cell on a Step tab (e.g., ='Step 2'!E12). "
         "Do not retype values into the Answer Sheet — link to your work."),
        ("Step 1 — Identify the cost driver",
         "The data dump contains five candidate cost drivers. For each, build a scatter plot against "
         "indirect labor cost and compute the correlation coefficient. Decide which driver best explains "
         "the cost — defend your choice using (a) the visual fit of the scatter plot, (b) the correlation, "
         "and (c) whether the relationship makes economic sense for this kind of business."),
        ("Step 2 — Estimate the cost function",
         "Using your chosen driver from Step 1, estimate a linear cost function Y = a + bX with regression. "
         "Report the fixed cost, variable cost, and R². Then estimate the same cost function with the "
         "high-low method. Compare and discuss why the two methods can give different answers."),
        ("Step 3 — Predict costs",
         "Use your Step 2 regression cost function to predict indirect labor at three batch volumes: "
         "75, 175, and 250. For each, comment on whether the prediction is reliable given the historical "
         "relevant range."),
        ("Step 4 — Plant expansion scenario",
         "Pearson is considering a plant expansion: monthly fixed indirect labor would rise by 20% (more "
         "supervisors, more facility overhead) and variable indirect labor would fall by 10% (better "
         "automation reduces per-batch sanitation labor). The new relevant range would be 125 to 300 "
         "batches per month. Plot the current and expanded cost functions on the same chart. Solve for "
         "the indifference batch volume Q*. Then make a recommendation: if next year's forecast is 240 "
         "batches/month, should Pearson expand?"),
        ("A note on Excel tools",
         "You will need: scatter plots (Insert → Chart → Scatter), CORREL or PEARSON for correlations, "
         "SLOPE / INTERCEPT / RSQ for regression (or the Data Analysis ToolPak's Regression tool), and "
         "MAX / MIN / INDEX / MATCH for the high-low method. Step 4 also asks you to plot two cost "
         "functions on one chart — the Step 4 tab walks you through how."),
        ("Submission",
         "Submit this workbook with the Answer Sheet tab completed. Grading reads only that tab, so "
         "every numeric answer must be a formula reference to a cell in your Step tab work."),
    ]

    r = 4
    for header, body in blocks:
        section(ws, r, header, span=10); r += 1
        c = ws.cell(row=r, column=2, value=body)
        c.font = F(); c.alignment = al_top
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        ws.row_dimensions[r].height = max(38, 14*max(1, len(body)//90))
        r += 2

# ---- Data tab ----
def build_data(ws):
    ws.column_dimensions['A'].width = 12
    for c, w in zip("BCDEFGHIJ", [12, 16, 16, 18, 18, 20, 4, 14, 18]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "Pearson Specialty Foods — Operating Data Dump (24 months)", span=7)
    sub = ws.cell(row=2, column=1, value="Source: Operations Data Warehouse, January 2024 – December 2025")
    sub.font = F(italic=True, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    # Header row
    headers = ['Month', 'Batches', 'Units (cases)', 'Machine Hours', 'Direct Labor Hours',
               'Raw Material (lbs)', 'Indirect Labor Cost']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF")
        c.fill = fill(NAVY)
        c.alignment = al_c
        c.border = box
    ws.row_dimensions[3].height = 32

    # Data rows
    for ri, row in DATA.iterrows():
        excel_row = DATA_FIRST_ROW + ri
        for ci, col in enumerate(headers, start=1):
            val = row[col]
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.font = F()
            c.alignment = al_c if ci == 1 else al_r
            c.border = box
            if col == 'Indirect Labor Cost':
                c.number_format = '"$"#,##0'
            elif col != 'Month':
                c.number_format = '#,##0'

    # Helpful labeled ranges note
    note_row = DATA_LAST_ROW + 2
    ws.cell(row=note_row, column=1,
            value="Tip: When building scatter plots and regressions on other tabs, reference cells on "
                  "this tab. For example, the batches column is 'Data'!B4:B27 and the indirect labor "
                  "cost column is 'Data'!G4:G27.").font = F(italic=True, color="595959")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
    ws.row_dimensions[note_row].height = 32


# ---- Step 1 ----
def build_step1(ws):
    ws.column_dimensions['A'].width = 5
    for c, w in zip("BCDEFGHIJK", [30, 18, 18, 5, 16, 16, 5, 16, 16, 16]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "Step 1 — Identify the Cost Driver", span=10)
    p = ws.cell(row=2, column=1,
                value="Build a scatter plot for each candidate driver against indirect labor cost. "
                      "Compute the correlation coefficient between each driver and the cost. Then make "
                      "your call: which driver best explains the variation in indirect labor cost?")
    p.font = F(); p.alignment = al_top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 42

    section(ws, 4, "1A — Correlation table", span=10)
    headers = ["Candidate cost driver", "Correlation (r) with IL", "R²"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=5, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF"); c.fill = fill(NAVY); c.alignment = al_c; c.border = box
    drivers = [
        ("Batches",            "Data!B4:B27"),
        ("Units (cases)",      "Data!C4:C27"),
        ("Machine Hours",      "Data!D4:D27"),
        ("Direct Labor Hours", "Data!E4:E27"),
        ("Raw Material (lbs)", "Data!F4:F27"),
    ]
    il_range = "Data!G4:G27"
    for i, (name, _rng) in enumerate(drivers):
        r = 6 + i
        ws.cell(row=r, column=2, value=name).font = F()
        ws.cell(row=r, column=2).border = box; ws.cell(row=r, column=2).alignment = al_l
        # students enter formulas here (yellow input cells)
        for col in (3, 4):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill(INPUT_YELLOW); cc.border = box; cc.font = F(bold=True, color="0000CC")
            cc.alignment = al_c
            cc.number_format = '0.0000'
        # Hint: pre-set placeholder text
    hint = ws.cell(row=11, column=2,
                   value="Use =CORREL(driver_range, IL_range). For R² use =RSQ(IL_range, driver_range), "
                         "or simply square the correlation. The IL range is Data!G4:G27.")
    hint.font = F(italic=True, color="595959"); hint.alignment = al_top
    ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=4)
    ws.row_dimensions[11].height = 30

    section(ws, 13, "1B — Scatter plots", span=10)
    p2 = ws.cell(row=14, column=1,
                 value="Insert one scatter plot for each candidate driver (X = driver, Y = indirect "
                       "labor cost). Place them anywhere on this tab. Use Insert → Chart → Scatter, "
                       "then drag your data ranges from the Data tab. Add chart titles and axis labels.")
    p2.font = F(); p2.alignment = al_top
    ws.merge_cells(start_row=14, start_column=1, end_row=14, end_column=10)
    ws.row_dimensions[14].height = 42

    # Suggested chart placement boxes — light cream blocks as visual placeholders
    chart_row = 16
    for i, (name, _) in enumerate(drivers):
        col = 2 + (i % 3) * 4
        row = chart_row + (i // 3) * 14
        ws.merge_cells(start_row=row, start_column=col, end_row=row+12, end_column=col+2)
        c = ws.cell(row=row, column=col, value=f"[ Insert scatter plot here:\n   X = {name}\n   Y = Indirect Labor Cost ]")
        c.fill = fill(LIGHT_NAVY); c.font = F(italic=True, color="595959"); c.alignment = al_c
        c.border = box

    decision_row = 16 + 28
    section(ws, decision_row, "1C — Your decision", span=10)
    ws.cell(row=decision_row+1, column=2, value="Best cost driver:").font = F(bold=True)
    dv = DataValidation(type="list",
                        formula1='"Batches,Units (cases),Machine Hours,Direct Labor Hours,Raw Material (lbs)"',
                        allow_blank=True)
    dv.error = "Pick one of the listed drivers."
    dv.errorTitle = "Invalid driver"
    ws.add_data_validation(dv)
    cell_addr = f"C{decision_row+1}"
    dv.add(cell_addr)
    cc = ws.cell(row=decision_row+1, column=3)
    cc.fill = fill(INPUT_YELLOW); cc.font = F(bold=True, color="0000CC"); cc.border = box; cc.alignment = al_c

    rationale = ws.cell(row=decision_row+3, column=2,
                        value="Defend your choice in 1–2 sentences (visual fit, correlation, economic logic):")
    rationale.font = F(italic=True, color="595959"); rationale.alignment = al_top
    ws.merge_cells(start_row=decision_row+3, start_column=2, end_row=decision_row+3, end_column=10)
    ws.merge_cells(start_row=decision_row+4, start_column=2, end_row=decision_row+7, end_column=10)
    ws.cell(row=decision_row+4, column=2).fill = fill(INPUT_YELLOW)
    ws.cell(row=decision_row+4, column=2).border = box
    ws.cell(row=decision_row+4, column=2).alignment = al_top


# ---- Step 2 ----
def build_step2(ws):
    ws.column_dimensions['A'].width = 5
    for c, w in zip("BCDEFGHI", [32, 18, 18, 5, 22, 18, 18, 18]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "Step 2 — Estimate the Cost Function", span=8)
    p = ws.cell(row=2, column=1,
                value="Using the cost driver you selected in Step 1, estimate a linear cost function "
                      "Y = a + bX two ways: with simple regression (SLOPE / INTERCEPT / RSQ), and with "
                      "the high-low method. Then compare them.")
    p.font = F(); p.alignment = al_top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 42

    # 2A regression
    section(ws, 4, "2A — Simple regression", span=8)
    ws.cell(row=5, column=2, value="Driver range (X):").font = F(bold=True)
    ws.cell(row=5, column=3).fill = fill(INPUT_YELLOW); ws.cell(row=5, column=3).border = box
    ws.cell(row=5, column=3).font = F(bold=True, color="0000CC")
    ws.cell(row=5, column=3).alignment = al_c
    ws.cell(row=5, column=4, value="(e.g., Data!B4:B27)").font = F(italic=True, color="595959")

    ws.cell(row=6, column=2, value="IL range (Y):").font = F(bold=True)
    ws.cell(row=6, column=3, value="Data!G4:G27").font = F()
    ws.cell(row=6, column=3).alignment = al_c
    ws.cell(row=6, column=3).border = box
    ws.cell(row=6, column=4, value="(this is fixed — IL is always Y)").font = F(italic=True, color="595959")

    headers2A = [
        ("Estimated variable cost (slope, b):", "$#,##0.0000"),
        ("Estimated fixed cost (intercept, a):", '"$"#,##0.00'),
        ("R² (coefficient of determination):", "0.0000"),
    ]
    for i, (lbl, fmt) in enumerate(headers2A):
        r = 8 + i
        ws.cell(row=r, column=2, value=lbl).font = F(bold=True)
        ws.cell(row=r, column=2).alignment = al_l
        cc = ws.cell(row=r, column=3)
        cc.fill = fill(INPUT_YELLOW); cc.border = box
        cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
        cc.number_format = fmt

    eq = ws.cell(row=12, column=2, value="Cost equation in plain English:")
    eq.font = F(bold=True)
    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=8)
    eqcell = ws.cell(row=13, column=2,
                     value="Indirect Labor = $______ + $______ × ______")
    eqcell.fill = fill(INPUT_YELLOW); eqcell.border = box
    eqcell.font = F(italic=True, color="0000CC"); eqcell.alignment = al_l

    hint2A = ws.cell(row=15, column=2,
                     value="Hints: =SLOPE(Y, X) gives b; =INTERCEPT(Y, X) gives a; =RSQ(Y, X) gives R². "
                           "Or run Data → Data Analysis → Regression for the full output.")
    hint2A.font = F(italic=True, color="595959"); hint2A.alignment = al_top
    ws.merge_cells(start_row=15, start_column=2, end_row=15, end_column=8)
    ws.row_dimensions[15].height = 30

    # 2B High-low
    section(ws, 17, "2B — High-low method", span=8)
    p2 = ws.cell(row=18, column=1,
                 value="Identify the months with the highest and lowest values of your chosen driver. "
                       "Use those two points to estimate variable cost and fixed cost.")
    p2.font = F(); p2.alignment = al_top
    ws.merge_cells(start_row=18, start_column=1, end_row=18, end_column=8)
    ws.row_dimensions[18].height = 30

    hl_headers = [("","Driver value (X)","IL value (Y)")]
    ws.cell(row=20, column=2, value="").font = F(bold=True)
    for i, h in enumerate(["Driver value (X)", "IL value (Y)"], start=3):
        c = ws.cell(row=20, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF"); c.fill = fill(NAVY); c.border = box; c.alignment = al_c
    for i, lbl in enumerate(["High month:", "Low month:"]):
        r = 21 + i
        ws.cell(row=r, column=2, value=lbl).font = F(bold=True)
        for col in (3, 4):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill(INPUT_YELLOW); cc.border = box
            cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
            cc.number_format = '#,##0' if col == 3 else '"$"#,##0'

    for i, (lbl, fmt) in enumerate([
        ("High-low variable cost (b):", "$#,##0.0000"),
        ("High-low fixed cost (a):", '"$"#,##0.00'),
    ]):
        r = 24 + i
        ws.cell(row=r, column=2, value=lbl).font = F(bold=True)
        cc = ws.cell(row=r, column=3)
        cc.fill = fill(INPUT_YELLOW); cc.border = box
        cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
        cc.number_format = fmt

    hint2B = ws.cell(row=27, column=2,
                     value="Hints: For the high/low values, =MAX(driver_range) and =MIN(driver_range) get the "
                           "X values; pair with =INDEX(Data!G4:G27, MATCH(MAX(driver), driver_range, 0)) "
                           "to pull the matching IL value. Then b = ΔY / ΔX, and a = Y − b·X at either anchor.")
    hint2B.font = F(italic=True, color="595959"); hint2B.alignment = al_top
    ws.merge_cells(start_row=27, start_column=2, end_row=27, end_column=8)
    ws.row_dimensions[27].height = 44

    # 2C compare
    section(ws, 29, "2C — Compare the two methods", span=8)
    p3 = ws.cell(row=30, column=1,
                 value="Why might regression and the high-low method give different answers? "
                       "(2–3 sentences. Hint: look closely at the November 2024 row in the data — what's "
                       "going on with that month?)")
    p3.font = F(); p3.alignment = al_top
    ws.merge_cells(start_row=30, start_column=1, end_row=30, end_column=8)
    ws.row_dimensions[30].height = 42
    ws.merge_cells(start_row=31, start_column=2, end_row=34, end_column=8)
    ws.cell(row=31, column=2).fill = fill(INPUT_YELLOW)
    ws.cell(row=31, column=2).border = box
    ws.cell(row=31, column=2).alignment = al_top


# ---- Step 3 ----
def build_step3(ws):
    ws.column_dimensions['A'].width = 5
    for c, w in zip("BCDEFGH", [22, 18, 22, 30, 30, 30]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "Step 3 — Predict Costs", span=7)
    p = ws.cell(row=2, column=1,
                value="Use your regression cost function from Step 2A to predict indirect labor cost at three "
                      "batch volumes: 75, 175, and 250. Reference the slope and intercept cells from Step 2 "
                      "rather than retyping the numbers.")
    p.font = F(); p.alignment = al_top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 42

    section(ws, 4, "3A — Predicted indirect labor cost", span=7)
    for i, h in enumerate(["Batches (X)", "Predicted IL (Y)"], start=2):
        c = ws.cell(row=5, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF"); c.fill = fill(NAVY); c.alignment = al_c; c.border = box
    for i, x in enumerate([75, 175, 250]):
        r = 6 + i
        ws.cell(row=r, column=2, value=x).font = F(); ws.cell(row=r, column=2).alignment = al_c
        ws.cell(row=r, column=2).border = box
        cc = ws.cell(row=r, column=3)
        cc.fill = fill(INPUT_YELLOW); cc.border = box
        cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
        cc.number_format = '"$"#,##0.00'

    hint = ws.cell(row=10, column=2,
                   value="Hint: Predicted Y = intercept + slope × X. Reference your Step 2 cells, e.g., "
                         "='Step 2'!C9 + 'Step 2'!C8 × 75.")
    hint.font = F(italic=True, color="595959"); hint.alignment = al_top
    ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=7)
    ws.row_dimensions[10].height = 30

    section(ws, 12, "3B — Reliability of the predictions", span=7)
    p2 = ws.cell(row=13, column=1,
                 value="The historical data on the Data tab spans batch volumes from roughly 35 to 190. "
                       "Which of your three predictions are inside that range, and which require "
                       "extrapolation? How confident should you be in each?")
    p2.font = F(); p2.alignment = al_top
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=7)
    ws.row_dimensions[13].height = 38
    ws.merge_cells(start_row=14, start_column=2, end_row=18, end_column=7)
    ws.cell(row=14, column=2).fill = fill(INPUT_YELLOW)
    ws.cell(row=14, column=2).border = box
    ws.cell(row=14, column=2).alignment = al_top


# ---- Step 4 ----
def build_step4(ws):
    ws.column_dimensions['A'].width = 5
    for c, w in zip("BCDEFGHIJK", [32, 18, 18, 5, 16, 18, 18, 5, 16, 16]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "Step 4 — Plant Expansion Scenario", span=10)
    p = ws.cell(row=2, column=1,
                value="Pearson is considering an expansion. Under the expansion: monthly fixed indirect "
                      "labor would rise by 20% (more supervisors, larger facility), and variable indirect "
                      "labor would fall by 10% (better automation reduces per-batch sanitation labor). "
                      "The new relevant range would be 125 to 300 batches per month.")
    p.font = F(); p.alignment = al_top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 56

    # 4A new cost function
    section(ws, 4, "4A — New cost function under expansion", span=10)
    rows4A = [
        ("New fixed cost (a' = 1.20 × Step 2 a):", '"$"#,##0.00'),
        ("New variable cost (b' = 0.90 × Step 2 b):", "$#,##0.0000"),
    ]
    for i, (lbl, fmt) in enumerate(rows4A):
        r = 5 + i
        ws.cell(row=r, column=2, value=lbl).font = F(bold=True)
        cc = ws.cell(row=r, column=3)
        cc.fill = fill(INPUT_YELLOW); cc.border = box
        cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
        cc.number_format = fmt

    # 4B chart-building data table
    section(ws, 8, "4B — Build the comparison chart", span=10)
    p4B = ws.cell(row=9, column=1,
                  value="Plot both cost functions (current and expanded) on the same scatter chart, with "
                        "batches on the x-axis and indirect labor cost on the y-axis. Use the new relevant "
                        "range (125 to 300) as your x-axis bounds. Steps:")
    p4B.font = F(); p4B.alignment = al_top
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=10)
    ws.row_dimensions[9].height = 38

    instructions = [
        "1. In the table below, fill the X column with batch values from 125 to 300 in steps of 25.",
        "2. Fill the 'Current cost' column using your Step 2A regression (= a + b × X).",
        "3. Fill the 'Expanded cost' column using your 4A new cost function (= a' + b' × X).",
        "4. Select all three columns including the headers; Insert → Chart → Scatter with Smooth Lines.",
        "5. Add a chart title (\"Current vs. Expanded Cost Functions\"), x-axis title (\"Batches per month\"), and y-axis title (\"Indirect Labor Cost ($)\"). The two lines should cross at Q*.",
    ]
    for i, t in enumerate(instructions):
        r = 10 + i
        ws.cell(row=r, column=2, value=t).font = F(); ws.cell(row=r, column=2).alignment = al_top
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        ws.row_dimensions[r].height = 18

    # Chart data table (students fill in)
    table_top = 16
    for i, h in enumerate(["Batches (X)", "Current cost", "Expanded cost"], start=2):
        c = ws.cell(row=table_top, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF"); c.fill = fill(NAVY); c.border = box; c.alignment = al_c
    for i, x in enumerate([125, 150, 175, 200, 225, 250, 275, 300]):
        r = table_top + 1 + i
        ws.cell(row=r, column=2, value=x).font = F(); ws.cell(row=r, column=2).alignment = al_c
        ws.cell(row=r, column=2).border = box
        for col in (3, 4):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill(INPUT_YELLOW); cc.border = box
            cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
            cc.number_format = '"$"#,##0'

    # 4C indifference Q*
    qstar_row = table_top + 11
    section(ws, qstar_row, "4C — Indifference batch volume Q*", span=10)
    explain = ws.cell(row=qstar_row+1, column=1,
                      value="Solve algebraically for Q*, the batch volume where current and expanded "
                            "total costs are equal. Set the two cost functions equal to each other and "
                            "solve for X:    a + bX = a' + b'X    ⟹    Q* = (a' − a) / (b − b')")
    explain.font = F(); explain.alignment = al_top
    ws.merge_cells(start_row=qstar_row+1, start_column=1, end_row=qstar_row+1, end_column=10)
    ws.row_dimensions[qstar_row+1].height = 42

    rows4C = [
        ("Indifference Q* (batches):", "#,##0.00"),
        ("Total cost at Q* (use either function — should match):", '"$"#,##0.00'),
    ]
    for i, (lbl, fmt) in enumerate(rows4C):
        r = qstar_row + 3 + i
        ws.cell(row=r, column=2, value=lbl).font = F(bold=True); ws.cell(row=r, column=2).alignment = al_l
        cc = ws.cell(row=r, column=3)
        cc.fill = fill(INPUT_YELLOW); cc.border = box
        cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
        cc.number_format = fmt

    # 4D recommendation
    rec_row = qstar_row + 7
    section(ws, rec_row, "4D — Recommendation at 240 batches/month", span=10)
    p4D = ws.cell(row=rec_row+1, column=1,
                  value="If next year's forecast is 240 batches/month, which cost structure is cheaper "
                        "and by how much?")
    p4D.font = F(); p4D.alignment = al_top
    ws.merge_cells(start_row=rec_row+1, start_column=1, end_row=rec_row+1, end_column=10)
    ws.row_dimensions[rec_row+1].height = 26

    rows4D = [
        ("Current cost at 240 batches:", '"$"#,##0.00'),
        ("Expanded cost at 240 batches:", '"$"#,##0.00'),
        ("Cheaper at 240:", "@"),
    ]
    for i, (lbl, fmt) in enumerate(rows4D):
        r = rec_row + 3 + i
        ws.cell(row=r, column=2, value=lbl).font = F(bold=True); ws.cell(row=r, column=2).alignment = al_l
        cc = ws.cell(row=r, column=3)
        cc.fill = fill(INPUT_YELLOW); cc.border = box
        cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
        cc.number_format = fmt

    dv = DataValidation(type="list", formula1='"Current,Expanded"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C{rec_row+5}")

    rec_text = ws.cell(row=rec_row+7, column=1,
                       value="In 1–3 sentences, what would you recommend Pearson do, and what additional "
                             "information would you want before making a final call?")
    rec_text.font = F(); rec_text.alignment = al_top
    ws.merge_cells(start_row=rec_row+7, start_column=1, end_row=rec_row+7, end_column=10)
    ws.row_dimensions[rec_row+7].height = 30
    ws.merge_cells(start_row=rec_row+8, start_column=2, end_row=rec_row+11, end_column=10)
    ws.cell(row=rec_row+8, column=2).fill = fill(INPUT_YELLOW)
    ws.cell(row=rec_row+8, column=2).border = box
    ws.cell(row=rec_row+8, column=2).alignment = al_top


# ---- Answer Sheet ----
QUESTIONS = [
    # (qid, text, ref_to_step_cell_for_instructor_key, format, points)
    ("1.1", "Best cost driver (text)", "='Step 1'!C45", "@", 10),
    ("1.2", "Correlation (r) with best driver", "='Step 1'!C6", "0.0000", 5),
    ("2.1", "Estimated variable cost b (slope)", "='Step 2'!C8", "$#,##0.0000", 8),
    ("2.2", "Estimated fixed cost a (intercept)", "='Step 2'!C9", '"$"#,##0.00', 8),
    ("2.3", "R² of the regression", "='Step 2'!C10", "0.0000", 5),
    ("2.4", "High-low variable cost", "='Step 2'!C24", "$#,##0.0000", 5),
    ("2.5", "High-low fixed cost", "='Step 2'!C25", '"$"#,##0.00', 5),
    ("3.1", "Predicted IL at 75 batches", "='Step 3'!C6", '"$"#,##0.00', 5),
    ("3.2", "Predicted IL at 175 batches", "='Step 3'!C7", '"$"#,##0.00', 5),
    ("3.3", "Predicted IL at 250 batches", "='Step 3'!C8", '"$"#,##0.00', 5),
    ("4.1", "New fixed cost (post-expansion)", "='Step 4'!C5", '"$"#,##0.00', 5),
    ("4.2", "New variable cost (post-expansion)", "='Step 4'!C6", "$#,##0.0000", 5),
    ("4.3", "Indifference Q* (batches)", "='Step 4'!C30", "#,##0.00", 14),
    ("4.4", "Total cost at Q*", "='Step 4'!C31", '"$"#,##0.00', 8),
    ("4.5", "Cheaper at 240 (Current/Expanded)", "='Step 4'!C39", "@", 7),
]

def build_answer_sheet(ws, is_instructor=False):
    ws.column_dimensions['A'].width = 5
    for c, w in zip("BCDE", [10, 50, 24, 10]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "Answer Sheet — link your final answers here", span=4)
    p = ws.cell(row=2, column=1,
                value=("Each cell in column D should be a FORMULA that links to the cell on a Step tab "
                       "where your answer lives. Example: ='Step 2'!C8 (do not retype the value). "
                       "Grading reads only this tab. Text answers (like the cost driver name in 1.1, "
                       "or Current/Expanded in 4.5) should also be linked from the Step tab."))
    p.font = F(); p.alignment = al_top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.row_dimensions[2].height = 56

    # header row
    headers = ["#", "Question", "Your Answer (link)", "Points"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF"); c.fill = fill(NAVY); c.alignment = al_c; c.border = box

    for i, (qid, text, ref, fmt, pts) in enumerate(QUESTIONS):
        r = 5 + i
        ws.cell(row=r, column=2, value=qid).font = F(bold=True)
        ws.cell(row=r, column=2).alignment = al_c; ws.cell(row=r, column=2).border = box
        ws.cell(row=r, column=3, value=text).font = F(); ws.cell(row=r, column=3).alignment = al_l
        ws.cell(row=r, column=3).border = box
        cc = ws.cell(row=r, column=4)
        cc.fill = fill(INPUT_YELLOW); cc.border = box
        cc.font = F(bold=True, color="0000CC"); cc.alignment = al_c
        cc.number_format = fmt
        ws.cell(row=r, column=5, value=pts).font = F(); ws.cell(row=r, column=5).alignment = al_c
        ws.cell(row=r, column=5).border = box
        if is_instructor:
            cc.value = ref  # populate with the reference for instructor key
            cc.font = F(bold=True, color="006600")  # green = link

    # Totals
    total_row = 5 + len(QUESTIONS) + 1
    ws.cell(row=total_row, column=3, value="TOTAL POSSIBLE POINTS").font = F(bold=True)
    ws.cell(row=total_row, column=3).alignment = al_r
    tc = ws.cell(row=total_row, column=5, value=f"=SUM(E5:E{4+len(QUESTIONS)})")
    tc.font = F(bold=True); tc.alignment = al_c; tc.border = box

# ===========================================================================
# INSTRUCTOR WORKBOOK — same as student + Grading tab + completed answer key
# ===========================================================================
def build_instructor(path):
    wb = Workbook()
    ws = wb.active; ws.title = "Instructions"
    build_instructions(ws)
    build_data(wb.create_sheet("Data"))
    build_step1(wb.create_sheet("Step 1"))
    build_step2(wb.create_sheet("Step 2"))
    build_step3(wb.create_sheet("Step 3"))
    build_step4(wb.create_sheet("Step 4"))
    build_answer_sheet(wb.create_sheet("Answer Sheet"), is_instructor=False)  # student-style; grading reads it
    build_answer_key(wb.create_sheet("ANSWER KEY"))
    build_grading(wb.create_sheet("Grading"))
    wb.save(path)
    print(f"Saved {path}")


# ---- ANSWER KEY tab ----
# Uses formulas referencing Data tab so that if data ever changes, key updates.
def build_answer_key(ws):
    ws.sheet_properties.tabColor = NAVY
    ws.column_dimensions['A'].width = 5
    for c, w in zip("BCDE", [10, 50, 22, 22]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "ANSWER KEY (instructor only)", span=4, color=ORANGE)
    note = ws.cell(row=2, column=1,
                   value="All values are computed via formulas off the Data tab. Edit the data and the "
                         "key will recalculate automatically. Grading tolerance bands live on the Grading tab.")
    note.font = F(italic=True, color="595959"); note.alignment = al_top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.row_dimensions[2].height = 32

    headers = ["#", "Question", "Expected Answer", "Tolerance"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF"); c.fill = fill(NAVY); c.alignment = al_c; c.border = box

    # Data ranges
    BX = "Data!B4:B27"   # batches
    UX = "Data!C4:C27"
    MX = "Data!D4:D27"
    DX = "Data!E4:E27"
    RX = "Data!F4:F27"
    Y  = "Data!G4:G27"

    # Formulas for each expected value
    expected = {
        "1.1": ("Batches", "@", "exact"),
        "1.2": (f"=CORREL({BX},{Y})", "0.0000", "±0.005"),
        "2.1": (f"=SLOPE({Y},{BX})", "$#,##0.0000", "±$2.00"),
        "2.2": (f"=INTERCEPT({Y},{BX})", '"$"#,##0.00', "±$300"),
        "2.3": (f"=RSQ({Y},{BX})", "0.0000", "±0.005"),
        "2.4": (f"=(INDEX({Y},MATCH(MAX({BX}),{BX},0))-INDEX({Y},MATCH(MIN({BX}),{BX},0)))/(MAX({BX})-MIN({BX}))",
                "$#,##0.0000", "±$2.00"),
        "2.5": (f"=INDEX({Y},MATCH(MAX({BX}),{BX},0))-((INDEX({Y},MATCH(MAX({BX}),{BX},0))-INDEX({Y},MATCH(MIN({BX}),{BX},0)))/(MAX({BX})-MIN({BX})))*MAX({BX})",
                '"$"#,##0.00', "±$300"),
        "3.1": (f"=INTERCEPT({Y},{BX})+SLOPE({Y},{BX})*75", '"$"#,##0.00', "±$300"),
        "3.2": (f"=INTERCEPT({Y},{BX})+SLOPE({Y},{BX})*175", '"$"#,##0.00', "±$300"),
        "3.3": (f"=INTERCEPT({Y},{BX})+SLOPE({Y},{BX})*250", '"$"#,##0.00', "±$300"),
        "4.1": (f"=1.20*INTERCEPT({Y},{BX})", '"$"#,##0.00', "±$300"),
        "4.2": (f"=0.90*SLOPE({Y},{BX})", "$#,##0.0000", "±$2.00"),
        "4.3": (f"=(1.20*INTERCEPT({Y},{BX})-INTERCEPT({Y},{BX}))/(SLOPE({Y},{BX})-0.90*SLOPE({Y},{BX}))",
                "#,##0.00", "±2.0"),
        "4.4": ("=INTERCEPT(Data!G4:G27,Data!B4:B27)+SLOPE(Data!G4:G27,Data!B4:B27)*"
                "((1.20*INTERCEPT(Data!G4:G27,Data!B4:B27)-INTERCEPT(Data!G4:G27,Data!B4:B27))/"
                "(SLOPE(Data!G4:G27,Data!B4:B27)-0.90*SLOPE(Data!G4:G27,Data!B4:B27)))",
                '"$"#,##0.00', "±$300"),
        "4.5": ("Expanded", "@", "exact"),
    }

    for i, (qid, text, _ref, _fmt, _pts) in enumerate(QUESTIONS):
        r = 5 + i
        formula, fmt, tol = expected[qid]
        ws.cell(row=r, column=2, value=qid).font = F(bold=True)
        ws.cell(row=r, column=2).alignment = al_c; ws.cell(row=r, column=2).border = box
        ws.cell(row=r, column=3, value=text).font = F(); ws.cell(row=r, column=3).alignment = al_l
        ws.cell(row=r, column=3).border = box
        cc = ws.cell(row=r, column=4, value=formula)
        cc.font = F(); cc.alignment = al_c; cc.border = box; cc.number_format = fmt
        tc = ws.cell(row=r, column=5, value=tol)
        tc.font = F(italic=True, color="595959"); tc.alignment = al_c; tc.border = box


# ---- Grading tab ----
# Compares Answer Sheet column D against ANSWER KEY column D using tolerance bands.
def build_grading(ws):
    ws.sheet_properties.tabColor = ORANGE
    ws.column_dimensions['A'].width = 5
    for c, w in zip("BCDEFGH", [8, 38, 18, 18, 14, 12, 12]):
        ws.column_dimensions[c].width = w

    banner(ws, 1, "Grading — auto-checks Answer Sheet against expected", span=7, color=ORANGE)
    note = ws.cell(row=2, column=1,
                   value="Drop a student-completed workbook on top of this template (or import their Answer Sheet "
                         "values) and this tab will score it automatically using tolerance bands from the "
                         "ANSWER KEY tab.")
    note.font = F(italic=True, color="595959"); note.alignment = al_top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 32

    headers = ["#", "Question", "Expected", "Student", "Diff", "Pass?", "Pts"]
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = F(bold=True, color="FFFFFF"); c.fill = fill(NAVY); c.alignment = al_c; c.border = box

    # Tolerance config (numeric absolute or relative). For text, exact match.
    tolerances = {
        "1.1": ("text", None),
        "1.2": ("abs", 0.005),
        "2.1": ("abs", 2.00),
        "2.2": ("abs", 300),
        "2.3": ("abs", 0.005),
        "2.4": ("abs", 2.00),
        "2.5": ("abs", 300),
        "3.1": ("abs", 300),
        "3.2": ("abs", 300),
        "3.3": ("abs", 300),
        "4.1": ("abs", 300),
        "4.2": ("abs", 2.00),
        "4.3": ("abs", 2.0),
        "4.4": ("abs", 300),
        "4.5": ("text", None),
    }

    fmt_map = {q: f for q, _, _, f, _ in [(q[0], None, None, q[3], None) for q in QUESTIONS]}
    pts_map = {q[0]: q[4] for q in QUESTIONS}

    for i, (qid, text, _ref, fmt, pts) in enumerate(QUESTIONS):
        r = 5 + i
        ws.cell(row=r, column=2, value=qid).font = F(bold=True); ws.cell(row=r, column=2).alignment = al_c
        ws.cell(row=r, column=2).border = box
        ws.cell(row=r, column=3, value=text).font = F(); ws.cell(row=r, column=3).alignment = al_l
        ws.cell(row=r, column=3).border = box

        # Expected from ANSWER KEY column D
        key_row = 5 + i  # parallel rows
        exp = ws.cell(row=r, column=4, value=f"='ANSWER KEY'!D{key_row}")
        exp.alignment = al_c; exp.border = box; exp.number_format = fmt

        # Student answer from Answer Sheet column D
        stu = ws.cell(row=r, column=5, value=f"='Answer Sheet'!D{4+i+1}")
        stu.alignment = al_c; stu.border = box; stu.number_format = fmt

        # Diff and Pass formulas
        kind, tol = tolerances[qid]
        if kind == "text":
            ws.cell(row=r, column=6, value="").alignment = al_c  # no numeric diff
            ws.cell(row=r, column=6).border = box
            pass_formula = (f'=IF(IFERROR(EXACT(TRIM(LOWER(D{r})),TRIM(LOWER(E{r}))),FALSE),"Pass","Fail")')
        else:
            ws.cell(row=r, column=6, value=f"=IFERROR(E{r}-D{r},NA())").alignment = al_c
            ws.cell(row=r, column=6).border = box
            ws.cell(row=r, column=6).number_format = fmt
            pass_formula = (f'=IF(IFERROR(ABS(E{r}-D{r})<={tol},FALSE),"Pass","Fail")')

        pc = ws.cell(row=r, column=7, value=pass_formula)
        pc.alignment = al_c; pc.border = box

        # Points awarded
        pp = ws.cell(row=r, column=8, value=f'=IF(G{r}="Pass",{pts},0)')
        pp.alignment = al_c; pp.border = box; pp.font = F(bold=True)

    # Totals row
    total_row = 5 + len(QUESTIONS)
    ws.cell(row=total_row+1, column=3, value="TOTAL").font = F(bold=True)
    ws.cell(row=total_row+1, column=3).alignment = al_r
    tot = ws.cell(row=total_row+1, column=8, value=f"=SUM(H5:H{total_row})")
    tot.font = F(bold=True, size=12); tot.fill = fill(CREAM); tot.alignment = al_c; tot.border = box

    pct = ws.cell(row=total_row+2, column=3, value="PERCENT")
    pct.font = F(bold=True); pct.alignment = al_r
    pctc = ws.cell(row=total_row+2, column=8, value=f"=H{total_row+1}/SUM(E5:E{total_row}*0+H5:H{total_row}*0+ROW(H5:H{total_row})*0+IF(TRUE,{{")
    # easier: just sum the maximum points
    pctc.value = f"=H{total_row+1}/{sum(q[4] for q in QUESTIONS)}"
    pctc.font = F(bold=True, size=12); pctc.fill = fill(CREAM); pctc.alignment = al_c
    pctc.border = box; pctc.number_format = "0.0%"

    # Conditional formatting: pass → green, fail → red
    from openpyxl.formatting.rule import FormulaRule, CellIsRule
    rng = f"G5:G{4+len(QUESTIONS)}"
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"Pass"'], fill=fill(GREEN_PASS)))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"Fail"'], fill=fill(RED_FAIL)))


# ---- main ----
build_student("AE1_Student.xlsx")
build_instructor("AE1_Instructor.xlsx")
print("Done.")
